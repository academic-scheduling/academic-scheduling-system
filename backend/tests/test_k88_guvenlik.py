"""K-88 · Güvenlik sertleştirmeleri — kapatılan açıkların regresyonu.

Bu dosya K-78'deki desenin devamı: bir güvenlik kararı yalnız kodda durursa
sessizce geri alınabilir, testte durursa alınamaz. Aşağıdaki her test, kapatılan
somut bir saldırıyı temsil eder.

Kapsam:
  1. Giriş kaba kuvvet freni (sınırsız parola denemesi)
  2. Giriş zamanlama sızıntısı (hesap sayımı)
  3. Bozuk JWT `sub` → 401, 500 değil
  4. CSV/XLSX formül enjeksiyonu
  5. Hoca import'unda SSRF kapısı
"""

import uuid

import jwt
import pytest

from app.config import settings
from app.db import SessionLocal
from app.export_service import formulu_kacir, to_csv_bytes, to_xlsx_bytes
from app.models import LoginAttempt
from app.scrapers.mu_akademik import adres_izinli_mi
from tests.helpers import ADMIN, client


@pytest.fixture(autouse=True)
def _giris_defterini_temizle():
    """Her testten önce ve sonra `login_attempts` boşaltılır.

    Sayaç IP başına da işlediği ve TestClient'ın tüm testlerde aynı istemci
    adresini kullandığı için, buradaki denemeler temizlenmezse paketin geri
    kalanındaki normal girişleri kilitleyebilirdi.
    """
    def temizle():
        db = SessionLocal()
        try:
            db.query(LoginAttempt).delete()
            db.commit()
        finally:
            db.close()

    temizle()
    yield
    temizle()


def _yanlis_giris(email: str):
    return client.post("/auth/login", json={"email": email, "password": "kesinlikle-yanlis"})


# ------------------------------------------------------------------
# 1. Kaba kuvvet freni
# ------------------------------------------------------------------

def test_ust_uste_yanlis_parola_sonunda_429_alir():
    """Sınırsız deneme kapandı: pencere içinde sınır dolunca 429 döner."""
    email = f"fren_{uuid.uuid4().hex[:8]}@muh.example.edu.tr"

    for _ in range(settings.login_max_failures_per_email):
        assert _yanlis_giris(email).status_code == 401

    assert _yanlis_giris(email).status_code == 429


def test_fren_dogru_parolayi_da_durdurur():
    """Fren parola denetiminden ÖNCE gelir.

    Aksi halde saldırgan doğru parolayı bulduğu anda frenden muaf olurdu; oysa
    frenin amacı tam olarak o anın gelmesini engellemek.
    """
    db = SessionLocal()
    try:
        from app.models import User, UserRole, UserStatus, Workgroup
        from app.security import hash_password

        email = f"fren2_{uuid.uuid4().hex[:8]}@muh.example.edu.tr"
        wg = db.query(Workgroup).first()
        db.add(User(
            workgroup_id=wg.id, name="Fren Denek", email=email,
            password_hash=hash_password("dogruparola123"),
            role=UserRole.SUB_ACCOUNT, status=UserStatus.ACTIVE,
        ))
        db.commit()
    finally:
        db.close()

    # Doğru parola önce çalışıyor.
    assert client.post(
        "/auth/login", json={"email": email, "password": "dogruparola123"}
    ).status_code == 200

    for _ in range(settings.login_max_failures_per_email):
        assert _yanlis_giris(email).status_code == 401

    # Sınır dolduktan sonra DOĞRU parola bile geçmez.
    assert client.post(
        "/auth/login", json={"email": email, "password": "dogruparola123"}
    ).status_code == 429


def test_basarisiz_deneme_deftere_yazilir_ama_eposta_ham_degil():
    """Sayaç var olmayan adresleri de sayar; ama adresi HAM saklamaz.

    Tablo kayıtsız adresleri de biriktirdiği için ham saklamak, sistemde hiç
    hesabı olmayan insanların adreslerinden bir liste üretirdi.
    """
    email = f"gizli_{uuid.uuid4().hex[:8]}@muh.example.edu.tr"
    _yanlis_giris(email)

    db = SessionLocal()
    try:
        kayitlar = db.query(LoginAttempt).all()
        assert len(kayitlar) == 1
        assert kayitlar[0].email_hash != email
        assert len(kayitlar[0].email_hash) == 64      # sha256 hex
        assert email not in kayitlar[0].email_hash
    finally:
        db.close()


def test_var_olmayan_adres_de_frenlenir():
    """429, hesabın var olduğunu ELE VERMEZ.

    Fren kayıtlı olmayan bir adresle de tetiklenebildiği için, 429 almak
    "bu adres sistemde var" anlamına gelmez.
    """
    yok = f"hicyok_{uuid.uuid4().hex[:8]}@muh.example.edu.tr"
    for _ in range(settings.login_max_failures_per_email):
        assert _yanlis_giris(yok).status_code == 401
    assert _yanlis_giris(yok).status_code == 429


def test_basarili_giris_defter_birakmaz():
    """Yalnız başarısızlıklar sayılır; normal kullanım kendini kilitlemez."""
    assert client.post("/auth/login", json=ADMIN).status_code == 200

    db = SessionLocal()
    try:
        assert db.query(LoginAttempt).count() == 0
    finally:
        db.close()


# ------------------------------------------------------------------
# 2. Zamanlama sızıntısı
# ------------------------------------------------------------------

def test_var_olmayan_hesapta_da_parola_dogrulamasi_kosar(monkeypatch):
    """Hesap yokken de bir bcrypt doğrulaması çalışır.

    Zamanlama farkı ölçerek test etmek gürültülü ve kararsız olurdu; bunun
    yerine `verify_password`'ün ÇAĞRILDIĞI doğrulanıyor — sızıntıyı kapatan
    şey tam olarak bu çağrının koşulsuz olması.
    """
    cagrildi: list[bool] = []
    gercek = __import__("app.routers.auth", fromlist=["x"]).verify_password

    def sayan(plain, hashed):
        cagrildi.append(True)
        return gercek(plain, hashed)

    monkeypatch.setattr("app.routers.auth.verify_password", sayan)

    yok = f"yok_{uuid.uuid4().hex[:8]}@muh.example.edu.tr"
    assert _yanlis_giris(yok).status_code == 401
    assert cagrildi, "hesap bulunamadığında parola doğrulaması atlanıyor (zamanlama sızıntısı)"


# ------------------------------------------------------------------
# 3. Bozuk JWT sub
# ------------------------------------------------------------------

@pytest.mark.parametrize("sub", ["abc", "", "1; DROP TABLE users"])
def test_sayiya_cevrilemeyen_sub_401_dondurur(sub):
    """Bozuk `sub` bir KİMLİK hatasıdır; ham int() çağrısı 500 üretiyordu."""
    token = jwt.encode({"sub": sub}, settings.secret_key, algorithm=settings.algorithm)
    r = client.get("/auth/me", headers={"Authorization": f"Bearer {token}"})
    assert r.status_code == 401


# ------------------------------------------------------------------
# 4. Formül enjeksiyonu
# ------------------------------------------------------------------

@pytest.mark.parametrize("kotu", [
    '=HYPERLINK("http://kotu","tikla")',
    "+1+1",
    "-2-3",
    "@SUM(A1:A9)",
])
def test_formul_baslangicli_metin_kacirilir(kotu):
    assert formulu_kacir(kotu) == "'" + kotu


def test_zararsiz_deger_bozulmaz():
    """Kaçış yalnız formül işaretine dokunur; sayılar ve normal metin aynı kalır."""
    assert formulu_kacir("CENG 101") == "CENG 101"
    assert formulu_kacir("Ali Veli") == "Ali Veli"
    assert formulu_kacir(42) == 42
    assert formulu_kacir(None) is None


def test_csv_ciktisinda_formul_calismaz():
    icerik = to_csv_bytes(["Ders"], [['=cmd|calc']]).decode("utf-8-sig")
    # Hücre "=" ile BAŞLAMAMALI; csv alıntılaması nedeniyle tırnak içinde olabilir.
    assert "'=cmd|calc" in icerik
    assert not any(
        h.lstrip('"').startswith("=") for h in icerik.splitlines()[1].split(",")
    )


def test_xlsx_hucresi_formul_olarak_yazilmaz():
    import io as _io

    from openpyxl import load_workbook

    ham = to_xlsx_bytes(["Ders"], [['=1+1']])
    ws = load_workbook(_io.BytesIO(ham)).active
    hucre = ws.cell(row=2, column=1)
    # data_type "f" olsaydi Excel bunu formul olarak calistirirdi.
    assert hucre.data_type != "f"
    assert str(hucre.value).startswith("'=") or not str(hucre.value).startswith("=")


# ------------------------------------------------------------------
# 5. SSRF kapısı
# ------------------------------------------------------------------

@pytest.mark.parametrize("url", [
    "http://169.254.169.254/latest/meta-data/",   # bulut metadata ucu
    "http://localhost:8000/health",
    "http://127.0.0.1/",
    "http://db:5432/",
    "file:///etc/passwd",
    "gopher://example.com/",
    "https://kotumu.edu.tr/personel/x",           # sona benzeyen yabanci alan
    "https://mu.edu.tr.kotu.com/personel/x",      # alan adi onek olarak gomulu
])
def test_izinsiz_adresler_reddedilir(url):
    assert adres_izinli_mi(url) is False


@pytest.mark.parametrize("url", [
    "https://muhendislik.mu.edu.tr/tr/personel/akademik",
    "https://www.mu.edu.tr/tr/personel/ornekkisi",
    "https://mu.edu.tr/tr/personel/ornekkisi",
])
def test_kurum_adresleri_kabul_edilir(url):
    """Liste ve detay farklı alt alanlarda; kısıt çalışan import'u kırmamalı."""
    assert adres_izinli_mi(url) is True
