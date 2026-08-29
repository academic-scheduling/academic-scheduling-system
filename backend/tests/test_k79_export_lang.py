"""K-79 Faz 4 · Export çıktıları dil seçimini takip eder.

Kullanıcı kararı: "export da dili takip etsin". Kapsam ayrımı önemli:
  - LİSTE çıktıları (CSV + düz XLSX) tamamen çevrilir: sütun başlıkları, gün
    adları, dönem/tür/sınav-türü etiketleri, dosya adı, sayfa adı.
  - RESMÎ ızgaralar (üniversite formatındaki sınav programı ve haftalık ızgara)
    K-09'da sabitlenmiş ŞABLONLARDIR ve zaten İngilizce başlıklıdır — dil
    düğmesi onların şablonunu DEĞİŞTİRMEZ. Buradaki test o sınırı da korur.
"""

import io

from openpyxl import load_workbook

from tests.helpers import admin_headers, client, publish_weekly, _u
from tests.test_k59_draft_api import (
    make_classroom, make_course, make_department, make_lecturer, make_section,
)


def _setup():
    """Bir bölüm + bir yayında haftalık yerleşim üretir."""
    h = admin_headers()
    dep = make_department(h)
    lec = make_lecturer(h)
    cls = make_classroom(h)
    course = make_course(h, dep["id"], year=1, semester="FALL")
    sec = make_section(h, course["id"], lec["id"])
    publish_weekly(sec["id"], classroom_id=cls["id"], day_of_week=1, start_slot=1)
    return h, dep


def _csv(path: str, headers: dict) -> str:
    r = client.get(path, headers=headers)
    assert r.status_code == 200, r.text
    return r.content.decode("utf-8-sig")


# ------------------------------------------------------------------
# Liste export'ları: başlıklar + etiketler çevrilir
# ------------------------------------------------------------------

def test_weekly_csv_headers_follow_language():
    h, dep = _setup()
    yol = f"/export/weekly?format=csv&department_id={dep['id']}"

    tr = _csv(yol, h)
    assert tr.startswith("Bölüm,")
    assert "Pazartesi" in tr

    en = _csv(yol, {**h, "Accept-Language": "en"})
    assert en.startswith("Department,")
    assert "Monday" in en
    assert "Pazartesi" not in en


def test_weekly_csv_translates_semester_and_session_labels():
    """Sütun başlığı yetmez: hücrelerdeki etiketler de çevrilmeli."""
    h, dep = _setup()
    en = _csv(f"/export/weekly?format=csv&department_id={dep['id']}",
              {**h, "Accept-Language": "en"})
    assert "Fall" in en and "Güz" not in en
    assert "Theory" in en and "Teori" not in en


def test_classroom_csv_headers_follow_language():
    h, _dep = _setup()
    tr = _csv("/export/classrooms?format=csv", h)
    assert tr.startswith("Derslik,")

    en = _csv("/export/classrooms?format=csv", {**h, "Accept-Language": "en"})
    assert en.startswith("Classroom,")


def test_exam_csv_headers_follow_language():
    h, dep = _setup()
    yol = f"/export/exams?format=csv&department_id={dep['id']}"
    tr = _csv(yol, h)
    assert tr.startswith("Bölüm,")

    en = _csv(yol, {**h, "Accept-Language": "en"})
    assert en.startswith("Department,")


def test_filename_follows_language():
    """Dosya adı da dili takip eder — indirilen dosya karışık dilde olmasın."""
    h, dep = _setup()
    yol = f"/export/weekly?format=csv&department_id={dep['id']}"

    tr = client.get(yol, headers=h)
    assert "haftalik_program.csv" in tr.headers["content-disposition"]

    en = client.get(yol, headers={**h, "Accept-Language": "en"})
    assert "weekly_schedule.csv" in en.headers["content-disposition"]


def test_classroom_xlsx_sheet_name_and_grid_headers_follow_language():
    h, _dep = _setup()
    r = client.get("/export/classrooms?format=xlsx",
                   headers={**h, "Accept-Language": "en"})
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws.title == "Classroom Schedule"

    duz = " ".join(
        str(v) for row in ws.iter_rows(values_only=True) for v in row if v
    )
    assert "Monday" in duz and "Pazartesi" not in duz
    assert "Time" in duz


# ------------------------------------------------------------------
# Sınır: resmî şablon dil düğmesinden ETKİLENMEZ
# ------------------------------------------------------------------

def test_official_exam_grid_template_is_not_translated():
    """Resmî sınav programı K-09 şablonudur: dil düğmesi şablonunu değiştirmez.

    Zaten İngilizce basılıyor (name_en/faculty_en + FIRST/SECOND YEAR);
    Türkçe istekte de İngilizce kalmalı — aksi hâlde kuruma giden belge
    kullanıcının arayüz tercihine göre değişirdi.
    """
    h, dep = _setup()
    r = client.get(f"/export/exams?format=xlsx&department_id={dep['id']}&semester=FALL",
                   headers=h)                     # Türkçe istek
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content)).active
    duz = " ".join(
        str(v) for row in ws.iter_rows(values_only=True) for v in row if v
    ).upper()
    # Şablonun İngilizce iskeleti Türkçe istekte de yerinde:
    assert "FALL" in duz or "SPRING" in duz
