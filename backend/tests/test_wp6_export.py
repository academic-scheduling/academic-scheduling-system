"""WP6 export API testleri — kontrat §11 + derslik programi.

Uc endpoint (weekly / exams / classrooms) x iki format (csv / xlsx). Dogrulanan:
- 200 + Content-Disposition attachment + dogru dosya adi/MIME
- CSV icerigi UTF-8 BOM'lu, baslik + bilinen kayit iceriyor
- XLSX gecerli (PK imzasi) ve openpyxl ile geri okunabiliyor
- Workgroup izolasyonu: yabanci admin bizim veriyi export'ta gormez (K-26)
- Sinav tip filtresi, desteklenmeyen format (400), auth (401)
"""

import io

from openpyxl import load_workbook

from tests.helpers import client, admin_headers, foreign_admin_headers, _u
from tests.helpers import publish_exam, publish_weekly

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# --- kurulum yardimcilari (wp3/wp4 desenleri) ---

def _post(h, path, body):
    r = client.post(path, json=body, headers=h)
    assert r.status_code == 201, r.text
    return r.json()


def _setup(h):
    """Bir bolum + ders + sube + DERSLIKLI haftalik giris + sinav kurar."""
    dep = _post(h, "/departments", {"name": "Export Bölüm", "code": _u("EB")})
    lec = _post(h, "/lecturers", {"full_name": f"Dr. Export {_u('')}"})
    building = _post(h, "/buildings", {"name": f"Export Bina {_u('')}"})
    room = _post(h, "/classrooms", {
        "building_id": building["id"], "room_code": _u("D"),
        "capacity": 90, "exam_capacity": 40,
    })
    course = _post(h, "/courses", {
        "department_id": dep["id"], "year": 2, "semester": "FALL",
        "code": _u("EX"), "name": "Export Ders",
        "hours_theory": 3, "hours_practice": 0, "hours_lab": 0,
    })
    section = _post(h, f"/courses/{course['id']}/sections", {
        "section_no": 1, "lecturer_id": lec["id"], "expected_students": 30,
    })
    publish_weekly(section["id"], classroom_id=room["id"],
                   day_of_week=1, start_slot=1, slot_count=2)
    publish_exam(course["id"], lec["id"], classroom_ids=[room["id"]])
    return {"dep": dep, "course": course, "room": room, "building": building, "lec": lec}


def _all_cell_text(content: bytes) -> str:
    """XLSX'teki tum hucre metnini tek stringde toplar (icerik aramasi icin)."""
    ws = load_workbook(io.BytesIO(content)).active
    out = []
    for row in ws.iter_rows(values_only=True):
        out += [str(v) for v in row if v is not None]
    return "\n".join(out)


# --- haftalik program ---

def test_weekly_csv():
    h = admin_headers()
    s = _setup(h)
    r = client.get("/export/weekly", params={"format": "csv"}, headers=h)
    assert r.status_code == 200, r.text
    cd = r.headers["content-disposition"]
    assert "attachment" in cd and "haftalik_program.csv" in cd
    assert r.headers["content-type"].startswith("text/csv")
    text = r.content.decode("utf-8-sig")
    assert text.startswith("Bölüm,")          # baslik satiri + BOM soyuldu
    assert s["course"]["code"] in text


def test_weekly_xlsx():
    h = admin_headers()
    s = _setup(h)
    r = client.get("/export/weekly", params={"format": "xlsx"}, headers=h)
    assert r.status_code == 200, r.text
    assert "haftalik_program.xlsx" in r.headers["content-disposition"]
    assert r.headers["content-type"].startswith(XLSX_MIME)
    assert r.content[:2] == b"PK"              # gecerli xlsx (zip)
    assert s["course"]["code"] in _all_cell_text(r.content)


def test_weekly_cohort_grid_xlsx():
    # Bolum + sinif + donem birlikte -> resmi IZGARA programi (duz liste degil).
    h = admin_headers()
    s = _setup(h)   # ders yil 2 GUZ, Pazartesi 1. slot haftalik giris
    r = client.get("/export/weekly", params={
        "format": "xlsx", "department_id": s["dep"]["id"],
        "year": 2, "semester": "FALL",
    }, headers=h)
    assert r.status_code == 200, r.text
    assert r.content[:2] == b"PK"
    body = _all_cell_text(r.content)
    assert "WEEKLY SCHEDULE" in body           # resmi izgara basligi
    assert "MONDAY" in body                     # gun basligi
    assert s["course"]["code"] in body          # ders hucresi


# --- sinav programi ---

def test_exams_midterm_schedule_xlsx():
    h = admin_headers()
    s = _setup(h)   # MIDTERM sinavi, ders yil 2 GUZ
    r = client.get("/export/exams", params={
        "format": "xlsx", "department_id": s["dep"]["id"],
        "semester": "FALL", "schedule": "midterm",
    }, headers=h)
    assert r.status_code == 200, r.text
    assert "vize_programi.xlsx" in r.headers["content-disposition"]
    assert r.content[:2] == b"PK"
    body = _all_cell_text(r.content)
    assert "MIDTERM EXAM SCHEDULE" in body      # ingilizce baslik
    assert "SECOND" in body                      # yil 2 -> SECOND grubu
    assert s["course"]["code"] in body


def test_exams_final_schedule_pairs_makeup():
    h = admin_headers()
    s = _setup(h)
    # Ayni derse FINAL + MAKEUP ekle.
    for typ, day in (("FINAL", "2026-11-12"), ("MAKEUP", "2026-11-19")):
        publish_exam(s["course"]["id"], s["lec"]["id"], exam_type=typ,
                     exam_date=day, classroom_ids=[s["room"]["id"]])

    r = client.get("/export/exams", params={
        "format": "xlsx", "department_id": s["dep"]["id"],
        "semester": "FALL", "schedule": "final",
    }, headers=h)
    assert r.status_code == 200, r.text
    assert "final_butunleme_programi.xlsx" in r.headers["content-disposition"]
    body = _all_cell_text(r.content)
    assert "FINAL AND MAKE UP EXAM SCHEDULE" in body
    assert "MAKE UP" in body                     # super-baslik blogu
    assert s["course"]["code"] in body

    # Vize programinda final/but GORUNMEZ (tur ayrimi).
    rm = client.get("/export/exams", params={
        "format": "xlsx", "department_id": s["dep"]["id"],
        "semester": "FALL", "schedule": "midterm",
    }, headers=h)
    mid_body = _all_cell_text(rm.content)
    assert "MIDTERM EXAM SCHEDULE" in mid_body   # vize var
    assert "MAKE UP" not in mid_body             # ama makeup blogu yok


def test_exams_header_uses_english_names():
    h = admin_headers()
    dep = _post(h, "/departments", {
        "name": "Bilgisayar Mühendisliği", "code": _u("CE"),
        "name_en": "Computer Engineering", "faculty_en": "Faculty of Engineering",
    })
    lec = _post(h, "/lecturers", {"full_name": f"Dr. {_u('')}"})
    course = _post(h, "/courses", {
        "department_id": dep["id"], "year": 1, "semester": "SPRING",
        "code": _u("CE"), "name": "Intro",
    })
    publish_exam(course["id"], lec["id"], exam_date="2026-04-20",
                 duration_minutes=60)
    r = client.get("/export/exams", params={
        "format": "xlsx", "department_id": dep["id"],
        "semester": "SPRING", "schedule": "midterm",
    }, headers=h)
    assert r.status_code == 200, r.text
    body = _all_cell_text(r.content)
    assert "FACULTY OF ENGINEERING" in body                  # fakülte satırı
    assert "DEPARTMENT OF COMPUTER ENGINEERING" in body      # ingilizce bölüm adı
    assert "FIRST" in body                                   # yıl 1 grubu


# --- derslik programi (izgara) ---

def test_classrooms_xlsx_grid():
    h = admin_headers()
    s = _setup(h)
    r = client.get("/export/classrooms", params={"format": "xlsx"}, headers=h)
    assert r.status_code == 200, r.text
    assert "derslik_programi.xlsx" in r.headers["content-disposition"]
    assert r.content[:2] == b"PK"
    body = _all_cell_text(r.content)
    assert s["building"]["name"] in body       # derslik basligi
    assert s["course"]["code"] in body         # dolu hucre
    assert "Pazartesi" in body                 # gun basligi


def test_classrooms_csv():
    h = admin_headers()
    s = _setup(h)
    r = client.get("/export/classrooms", params={"format": "csv"}, headers=h)
    assert r.status_code == 200, r.text
    assert "derslik_programi.csv" in r.headers["content-disposition"]
    assert s["course"]["code"] in r.content.decode("utf-8-sig")


# --- izolasyon / hata / auth ---

def test_isolation_foreign_admin_sees_nothing():
    h = admin_headers()
    s = _setup(h)
    fh = foreign_admin_headers()
    for path in ("/export/weekly", "/export/exams", "/export/classrooms"):
        r = client.get(path, params={"format": "csv"}, headers=fh)
        assert r.status_code == 200, r.text
        assert s["course"]["code"] not in r.content.decode("utf-8-sig"), path


def test_unsupported_format_400():
    h = admin_headers()
    r = client.get("/export/weekly", params={"format": "pdf"}, headers=h)
    assert r.status_code == 400


def test_requires_auth_401():
    r = client.get("/export/weekly", params={"format": "csv"})
    assert r.status_code == 401
