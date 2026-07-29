"""Export bicimleme servisi (WP6).

Router ince kalir; kayit -> bytes donusumu burada toplanir. Boylece
3 endpoint x 2 format tek yerden beslenir, tekrar olmaz (brief §10.2).

Slot saatleri ve gun adlari frontend/src/utils/slots.ts ile AYNIDIR;
orasi tek kaynak, burasi Python yansimasi (ayri dil oldugu icin kopya).
Slot tablosu degisirse iki yer birlikte guncellenir.
"""

import csv
import io
from collections import defaultdict
from datetime import date as _date, datetime as _datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- slots.ts yansimasi (kaynak: frontend/src/utils/slots.ts) ---
DAY_FULL = {1: "Pazartesi", 2: "Salı", 3: "Çarşamba", 4: "Perşembe", 5: "Cuma"}
SLOT_TIMES = {
    1: ("08:30", "09:15"), 2: ("09:30", "10:15"), 3: ("10:30", "11:15"),
    4: ("11:30", "12:15"), 5: ("12:30", "13:15"), 6: ("13:30", "14:15"),
    7: ("14:30", "15:15"), 8: ("15:30", "16:15"), 9: ("16:30", "17:15"),
}
MAX_SLOT = max(SLOT_TIMES)
SEMESTER_TR = {"FALL": "Güz", "SPRING": "Bahar", "SUMMER": "Yaz"}
SESSION_TR = {"THEORY": "Teori", "PRACTICE": "Uygulama", "LAB": "Lab"}
EXAMTYPE_TR = {"MIDTERM": "Vize", "FINAL": "Final", "MAKEUP": "Bütünleme"}


def slot_range(start_slot: int, slot_count: int) -> str:
    """'10:30 - 12:15' — cok slotlu ders son slotun bitisinde biter (slots.ts)."""
    start = SLOT_TIMES.get(start_slot)
    if not start:
        return ""
    end_slot = min(start_slot + slot_count - 1, MAX_SLOT)
    end = SLOT_TIMES.get(end_slot, start)
    return f"{start[0]} - {end[1]}"


def _classroom_label(entry) -> str:
    c = entry.classroom
    return f"{c.building.name} {c.room_code}" if c else "Çevrimiçi"


WEEKLY_HEADERS = [
    "Bölüm", "Yıl", "Dönem", "Ders Kodu", "Ders Adı", "Şube",
    "Öğretim Üyesi", "Gün", "Saat", "Derslik", "Tür", "Durum",
]


def weekly_rows(entries) -> list[list]:
    """Haftalik girisler -> duz satirlar (CSV ve XLSX liste sayfasi ortak)."""
    rows = []
    for e in entries:
        course = e.section.course
        rows.append([
            course.department.name,
            course.year,
            SEMESTER_TR.get(course.semester.value, course.semester.value),
            course.code,
            course.name,
            e.section.section_no,
            e.section.lecturer.full_name,
            DAY_FULL.get(e.day_of_week, "?"),
            slot_range(e.start_slot, e.slot_count),
            _classroom_label(e),
            SESSION_TR.get(e.session_type.value, e.session_type.value),
            e.status.value,
        ])
    return rows


def to_csv_bytes(headers: list[str], rows: list[list]) -> bytes:
    """UTF-8 BOM'lu CSV — Excel Turkce karakterleri dogru acsin diye."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")



def to_xlsx_bytes(headers: list[str], rows: list[list], sheet_name: str = "Program") -> bytes:
    """Duz tablo -> stilli tek sayfa XLSX. (headers, rows) alan her endpoint kullanir."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sayfa adi en fazla 31 karakter

    ws.append(headers)
    for cell in ws[1]:  # baslik satiri: kalin + ortali
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"  # kaydirinca baslik sabit kalir

    # Sutun genisligini basliga ve en uzun hucreye gore kabaca ayarla
    for i, header in enumerate(headers, start=1):
        longest = max([len(str(header))] + [len(str(r[i - 1])) for r in rows])
        ws.column_dimensions[get_column_letter(i)].width = min(longest + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _rooms_label(rooms) -> str:
    """Sinav coklu derslik olabilir (K-17); virgulle birlestir, bossa '—'."""
    return ", ".join(f"{r.building.name} {r.room_code}" for r in rooms) or "—"


EXAM_HEADERS = [
    "Bölüm", "Yıl", "Dönem", "Ders Kodu", "Ders Adı", "Sınav Türü",
    "Tarih", "Başlangıç", "Süre (dk)", "Öğretim Üyesi", "Derslikler",
    "Öğrenci", "Durum",
]


def exams_rows(exams) -> list[list]:
    """Sinav kayitlari -> duz satirlar (CSV + XLSX ortak)."""
    rows = []
    for x in exams:
        course = x.course
        rows.append([
            course.department.name,
            course.year,
            SEMESTER_TR.get(course.semester.value, course.semester.value),
            course.code,
            course.name,
            EXAMTYPE_TR.get(x.exam_type.value, x.exam_type.value),
            x.exam_date.strftime("%d.%m.%Y"),
            x.start_time.strftime("%H:%M"),
            x.duration_minutes,
            x.lecturer.full_name,
            _rooms_label(x.classrooms),
            x.total_expected_students,
            x.status.value,
        ])
    return rows    


DAYS = [1, 2, 3, 4, 5]  # Pzt–Cum (DAY_FULL anahtarlari)


def _cell_text(cell_entries) -> str:
    """Bir slottaki ders(ler): kod + ad + hoca, cok satirli."""
    parts = []
    for e in cell_entries:
        c = e.section.course
        parts.append(
            f"{c.code} · Şube {e.section.section_no}\n"
            f"{c.name}\n{e.section.lecturer.full_name}"
        )
    return "\n––\n".join(parts)  # ayni slotta birden fazla ders olursa ayirici


def build_classrooms_xlsx(entries) -> bytes:
    """Her derslik icin ayri izgara (UI ile ayni yon: satir=slot, sutun=gun).

    Ust satir derslik adi; dolu slotta ders kodu+adi+hoca. Cevrimici girisler
    (derslik yok) zaten sorguda inner join ile elenmis olarak gelir.
    Cok slotlu ders baslangic slotunda gorunur.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Derslik Programı"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=14)
    head_font = Font(bold=True)
    head_fill = PatternFill(fill_type="solid", fgColor="F0F0F0")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    by_room = defaultdict(list)
    for e in entries:
        by_room[e.classroom].append(e)

    row = 1
    for classroom in sorted(by_room, key=lambda c: (c.building.name, c.room_code)):
        # 1) Baslik: derslik adi, 6 sutuna yayilmis
        title = ws.cell(row=row, column=1,
                        value=f"{classroom.building.name} — {classroom.room_code}")
        title.font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

        # 2) Sutun basliklari: Saat + 5 gun
        for col, text in enumerate(["Saat"] + [DAY_FULL[d] for d in DAYS], start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font, cell.fill, cell.alignment, cell.border = head_font, head_fill, center, border
        row += 1

        # 3) Bu dersligin (gun, slot) -> girisler indeksi
        occ = defaultdict(list)
        for e in by_room[classroom]:
            # Cok slotlu ders isgal ettigi HER slotta ayri ayri yazilir
            # (kullanici tercihi); son slot gun sonunda kirpilir.
            last_slot = min(e.start_slot + e.slot_count - 1, MAX_SLOT)
            for s in range(e.start_slot, last_slot + 1):
                occ[(e.day_of_week, s)].append(e)

        # 4) 9 slot satiri
        for slot in range(1, MAX_SLOT + 1):
            start, end = SLOT_TIMES[slot]
            tcell = ws.cell(row=row, column=1, value=f"{start}–{end}")
            tcell.font, tcell.alignment, tcell.border = head_font, center, border
            for i, day in enumerate(DAYS):
                cell = ws.cell(row=row, column=2 + i)
                cell.alignment, cell.border = center, border
                if (day, slot) in occ:
                    cell.value = _cell_text(occ[(day, slot)])
            row += 1

        row += 1  # bloklar arasi bos satir

    ws.column_dimensions["A"].width = 14
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


CLASSROOM_HEADERS = [
    "Derslik", "Gün", "Saat", "Ders Kodu", "Ders Adı",
    "Şube", "Öğretim Üyesi", "Tür",
]


def classrooms_rows(entries) -> list[list]:
    """Derslik programi CSV (minimum): derslik -> gun -> saat sirali duz liste."""
    rows = []
    for e in sorted(entries, key=lambda e: (
        e.classroom.building.name, e.classroom.room_code,
        e.day_of_week, e.start_slot,
    )):
        c = e.section.course
        rows.append([
            f"{e.classroom.building.name} {e.classroom.room_code}",
            DAY_FULL.get(e.day_of_week, "?"),
            slot_range(e.start_slot, e.slot_count),
            c.code,
            c.name,
            e.section.section_no,
            e.section.lecturer.full_name,
            SESSION_TR.get(e.session_type.value, e.session_type.value),
        ])
    return rows

# ------------------------------------------------------------------
# Universite formatinda sinav programi (yila gore gruplu, resmi cikti)
# ------------------------------------------------------------------

YEAR_ORDINAL_EN = {1: "FIRST", 2: "SECOND", 3: "THIRD", 4: "FOURTH", 5: "FIFTH", 6: "SIXTH"}
SEASON_EN = {"FALL": "FALL", "SPRING": "SPRING", "SUMMER": "SUMMER"}


def _exam_date(exam) -> str:
    return exam.exam_date.strftime("%d.%m.%Y")


def _exam_time_range(exam) -> str:
    """'16:30 - 17:30' — baslangic + sure'den bitisi hesaplar."""
    end = (_datetime.combine(_date.min, exam.start_time)
           + timedelta(minutes=exam.duration_minutes)).time()
    return f"{exam.start_time.strftime('%H:%M')} - {end.strftime('%H:%M')}"


def _academic_year(exams, semester_value: str) -> str:
    """Sinav tarihlerinden akademik yil: bahar Y -> (Y-1)-Y, guz Y -> Y-(Y+1)."""
    dates = [e.exam_date for e in exams if e.exam_date]
    if not dates:
        return ""
    y = min(dates).year
    return f"{y - 1}-{y}" if semester_value == "SPRING" else f"{y}-{y + 1}"


def build_exam_schedule_xlsx(exams, *, faculty_en: str, department_en: str,
                             semester_value: str, schedule: str) -> bytes:
    """Universite RESMI sinav programi — orijinal PDF ile ayni yapi.

    2 satirlik baslik (akademik yil + sezon + FAKULTE / DEPARTMENT OF ... + tur),
    yila gore gruplar (FIRST/SECOND/THIRD/FORTH), iki satirlik sutun basligi
    (No/Kod/Ad/Hoca dikey birlesik + ustte MIDTERM ya da FINAL / MAKE UP).
    schedule='final': FINAL + MAKE UP iki blok, ders bazinda eslenir.
    """
    is_final = schedule == "final"
    ncols = 10 if is_final else 7
    wb = Workbook()
    ws = wb.active
    ws.title = "Final ve Butunleme" if is_final else "Vize"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold = Font(bold=True, size=9)

    def _border_range(r1, c1, r2, c2):
        # openpyxl: birlestirilmis hucrede kenarlik her hucreye ayri verilmeli.
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = border

    def merged(r1, c1, r2, c2, value, *, font=bold, align=center):
        ws.cell(row=r1, column=c1, value=value)
        if (r1, c1) != (r2, c2):
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        top = ws.cell(row=r1, column=c1)
        top.font = font
        top.alignment = align
        _border_range(r1, c1, r2, c2)

    def data_row(r, values):
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            cell.font = Font(size=9)
            text_cols = {3, 4, 7, 10} if is_final else {3, 4, 7}
            cell.alignment = left if i in text_cols else center
        return r + 1

    # --- Baslik: 2 satir, tam genislik (orijinaldeki gibi) ---
    ay = _academic_year(exams, semester_value)
    season = SEASON_EN.get(semester_value, "")
    title = "FINAL AND MAKE UP EXAM SCHEDULE" if is_final else "MIDTERM EXAM SCHEDULE"
    line1 = " ".join(x for x in [ay, season, "SEMESTER", faculty_en.upper()] if x)
    line2 = " ".join(x for x in [f"DEPARTMENT OF {department_en.upper()}".strip(), title] if x)
    merged(1, 1, 1, ncols, line1)
    merged(2, 1, 2, ncols, line2)
    row = 3

    by_year = defaultdict(list)
    for e in exams:
        by_year[e.course.year].append(e)

    for year in sorted(by_year):
        merged(row, 1, row, ncols, YEAR_ORDINAL_EN.get(year, f"{year}."))
        row += 1

        # Iki satirlik sutun basligi
        hr1, hr2 = row, row + 1
        for i, h in enumerate(["No", "Course Code", "Course Name", "Instructor"], start=1):
            merged(hr1, i, hr2, i, h)
        if is_final:
            merged(hr1, 5, hr1, 7, "FINAL")
            merged(hr1, 8, hr1, 10, "MAKE UP")
            for j, h in enumerate(["Date", "Time", "Classroom"] * 2, start=5):
                merged(hr2, j, hr2, j, h)
        else:
            merged(hr1, 5, hr1, 7, "MIDTERM")
            for j, h in enumerate(["Date", "Time", "Classroom"], start=5):
                merged(hr2, j, hr2, j, h)
        row = hr2 + 1

        if is_final:
            by_course: dict = {}
            for e in by_year[year]:
                item = by_course.setdefault(
                    e.course.id, {"course": e.course, "final": None, "makeup": None})
                if e.exam_type.value == "FINAL":
                    item["final"] = e
                elif e.exam_type.value == "MAKEUP":
                    item["makeup"] = e
            for n, it in enumerate(sorted(by_course.values(), key=lambda x: x["course"].code), start=1):
                c, f_e, m_e = it["course"], it["final"], it["makeup"]
                instr = (f_e or m_e).lecturer.full_name if (f_e or m_e) else ""
                row = data_row(row, [
                    n, c.code, c.name, instr,
                    _exam_date(f_e) if f_e else "", _exam_time_range(f_e) if f_e else "",
                    _rooms_label(f_e.classrooms) if f_e else "",
                    _exam_date(m_e) if m_e else "", _exam_time_range(m_e) if m_e else "",
                    _rooms_label(m_e.classrooms) if m_e else "",
                ])
        else:
            for n, e in enumerate(sorted(by_year[year], key=lambda e: e.course.code), start=1):
                row = data_row(row, [
                    n, e.course.code, e.course.name, e.lecturer.full_name,
                    _exam_date(e), _exam_time_range(e), _rooms_label(e.classrooms),
                ])

    widths = ([5, 13, 26, 30, 12, 13, 20, 12, 13, 20] if is_final
              else [5, 14, 34, 34, 13, 15, 26])
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------
# Haftalik program IZGARA export'u (universite formati, tek cohort)
# ------------------------------------------------------------------

WEEK_DAYS_EN = {1: "MONDAY", 2: "TUESDAY", 3: "WEDNESDAY", 4: "THURSDAY", 5: "FRIDAY"}
DAY_HEADER_FILL = {1: "FFFF00", 2: "C27BA0", 3: "00FFFF", 4: "F6B26B", 5: "FF00FF"}
GRID_START_MIN = 8 * 60 + 30     # 08:30
GRID_ROWS = 36                    # 08:30..17:30, 15 dk'lik satirlar
# Ders koduna gore deterministik renk paleti (pastel tonlar).
_WEEK_PALETTE = [
    "9FC5E8", "FFF2CC", "D0CECE", "FFC000", "F4CCCC", "2E75B5",
    "B6D7A8", "D9D2E9", "F9CB9C", "C9DAF8", "EAD1DC", "A2C4C9",
]


def _hhmm(mins: int) -> str:
    return f"{mins // 60:02d}:{mins % 60:02d}"


def _slot_minutes(entry):
    """Girisin baslangic ve bitis dakikasi (SLOT_TIMES'tan)."""
    start = SLOT_TIMES.get(entry.start_slot)
    last = min(entry.start_slot + entry.slot_count - 1, MAX_SLOT)
    end = SLOT_TIMES.get(last, start)
    to_min = lambda s: int(s[:2]) * 60 + int(s[3:])
    return to_min(start[0]), to_min(end[1])


def _week_color(code: str) -> str:
    import hashlib
    h = int(hashlib.md5(code.encode("utf-8")).hexdigest(), 16)
    return _WEEK_PALETTE[h % len(_WEEK_PALETTE)]


def _assign_lanes(day_entries):
    """Ayni gunde cakisan girisleri alt-sutunlara (lane) dagitir (interval boyama)."""
    lanes: list[int] = []          # her lane'in bittigi dakika
    placed = []
    for e in sorted(day_entries, key=lambda e: _slot_minutes(e)[0]):
        s, en = _slot_minutes(e)
        lane = next((i for i, end in enumerate(lanes) if end <= s), None)
        if lane is None:
            lane = len(lanes)
            lanes.append(en)
        else:
            lanes[lane] = en
        placed.append((e, lane))
    return placed, max(1, len(lanes))


def build_weekly_grid_xlsx(entries, *, faculty_en: str, department_en: str,
                           semester_value: str, year: int) -> bytes:
    """Tek cohort (bolum+yil+donem) haftalik programi — orijinal Excel ile ayni yapi.

    15 dk'lik satir izgarasi (08:30-17:30), sol tarafta Morning/BREAK/Afternoon +
    saat araliklari, ustte MONDAY-FRIDAY (renkli), her gun paralel sube sayisi kadar
    alt-sutun. Ders hucresi dikey birlesik, koda gore renkli.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{YEAR_ORDINAL_EN.get(year, str(year))} YEAR"[:31]

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True, size=9)

    # Gunlere gore grupla + lane ata
    by_day = defaultdict(list)
    for e in entries:
        by_day[e.day_of_week].append(e)
    lanes_of = {}
    placed_of = {}
    for d in range(1, 6):
        placed_of[d], lanes_of[d] = _assign_lanes(by_day.get(d, []))

    # Gun baslangic sutunlari (C=3'ten itibaren)
    day_start = {}
    col = 3
    for d in range(1, 6):
        day_start[d] = col
        col += lanes_of[d]
    ncols = col - 1                      # son dolu sutun
    last_col = get_column_letter(ncols)

    def fill(hexrgb):
        return PatternFill(fill_type="solid", fgColor="FF" + hexrgb)

    def _brange(r1, c1, r2, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = border

    def merged(r1, c1, r2, c2, value, *, font=bold, patternfill=None):
        ws.cell(row=r1, column=c1, value=value)
        if (r1, c1) != (r2, c2):
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1)
        cell.font = font
        cell.alignment = center
        if patternfill:
            cell.fill = patternfill
        _brange(r1, c1, r2, c2)

    # --- Baslik (satir 1) ---
    ay = _current_academic_year(semester_value)
    season = SEASON_EN.get(semester_value, "")
    title = " ".join(x for x in [
        ay, season, "SEMESTER", faculty_en.upper(),
        department_en.upper(), f"{YEAR_ORDINAL_EN.get(year, str(year))} YEAR WEEKLY SCHEDULE",
    ] if x)
    merged(1, 1, 1, ncols, title, font=Font(bold=True, size=12))

    # --- Satir 2: HOURS + gun basliklari ---
    merged(2, 1, 2, 2, "HOURS")
    for d in range(1, 6):
        c1 = day_start[d]
        # Renk YOK (kullanici istegi): gun basliklari beyaz, kalin + kenarlik.
        merged(2, c1, 2, c1 + lanes_of[d] - 1, WEEK_DAYS_EN[d],
               patternfill=fill("FFFFFF"))

    # --- Sol taraf: period etiketleri + saat araliklari (15 dk) ---
    # Period bloklari: Morning 08:30-12:30, BREAK 12:30-13:30, Afternoon 13:30-17:30
    def row_of(minute):
        return 3 + (minute - GRID_START_MIN) // 15
    rot = Alignment(horizontal="center", vertical="center", text_rotation=90)
    for r1r, r2r, label in [(row_of(510), row_of(750) - 1, "Morning"),   # 08:30-12:30
                            (row_of(750), row_of(810) - 1, "BREAK"),      # 12:30-13:30
                            (row_of(810), 2 + GRID_ROWS, "Afternoon")]:   # 13:30-17:30
        merged(r1r, 1, r2r, 1, label)
        ws.cell(row=r1r, column=1).alignment = rot                       # yan (dik) yazi
    for i in range(GRID_ROWS):
        r = 3 + i
        m = GRID_START_MIN + i * 15
        c = ws.cell(row=r, column=2, value=f"{_hhmm(m)} - {_hhmm(m + 15)}")
        c.font = Font(size=8)
        c.alignment = center
        c.border = border
        # Bos ders hucreleri BEYAZ: teneffusler ve bos slotlar beyaz gorunsun.
        for cc in range(3, ncols + 1):
            cell = ws.cell(row=r, column=cc)
            cell.border = border
            cell.fill = fill("FFFFFF")

    # --- Ders hucreleri: her 45 dk'lik slot AYRI renkli blok; aradaki teneffusler
    #     beyaz kalir (orijinaldeki gibi). Cok slotlu ders her slotta tekrar yazilir.
    def _min(s):
        return int(s[:2]) * 60 + int(s[3:])
    for d in range(1, 6):
        for e, lane in placed_of[d]:
            col = day_start[d] + lane
            sec = e.section
            crs = sec.course
            room = e.classroom
            lines = [f"{crs.code} - Section {sec.section_no}", crs.name, sec.lecturer.full_name]
            if room:
                lines.append(f"{room.building.name} {room.room_code}")
            text = "\n".join(lines)
            last = min(e.start_slot + e.slot_count - 1, MAX_SLOT)
            for slot in range(e.start_slot, last + 1):
                r1 = row_of(_min(SLOT_TIMES[slot][0]))
                # Renk YOK: ders hucreleri beyaz, yalniz kenarlik + metin (kullanici istegi).
                merged(r1, col, r1 + 2, col, text, font=Font(size=8),
                       patternfill=fill("FFFFFF"))

    # Sutun genislikleri (period etiketleri DIK yazildigindan A dar kalir)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 13
    for c in range(3, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _current_academic_year(semester_value: str) -> str:
    """Haftalik girislerde tarih yoktur; akademik yili GUNCEL takvimden turetir.
    Guz: Eylul-Aralik -> Y-(Y+1), Ocak-Agustos -> (Y-1)-Y. Bahar: (Y-1)-Y."""
    from datetime import date as _d
    t = _d.today()
    y = t.year
    if semester_value == "FALL":
        return f"{y}-{y + 1}" if t.month >= 9 else f"{y - 1}-{y}"
    return f"{y - 1}-{y}" if t.month <= 8 else f"{y}-{y + 1}"
