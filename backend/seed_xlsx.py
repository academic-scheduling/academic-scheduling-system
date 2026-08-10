"""GEÇİCİ test verisi: CENG 2025-2026 Bahar haftalık programını sisteme yükler.

KAYNAK: Desktop/staj_proje/CENG_2025_2026_Spring_Weekly_Schedule.xlsx
  - "1st..4th Year" sayfaları: 15 dk'lık grid, ders bloğu 3 satır (kod / ad / hoca)
  - "Common Courses" sayfası: tek hücrede çok satır (kod / ad / hoca / derslik)

ÇALIŞTIRMA (backend/ klasöründen, docker db ayaktayken):
    python seed_xlsx.py

Bu GEÇİCİ bir test aracıdır (seed_demo.py'nin gerçek-veri muadili); C'nin resmî
seed'i (docs/seed_data_plani.md) değildir. Giriş: admin@muh.example.edu.tr / admin1234

Dosyada OLMAYAN ve varsayılan atadığımız alanlar aşağıda "SENTETİK" diye işaretli.
"""

import re
from datetime import date, time
from collections import Counter, defaultdict

import openpyxl
from sqlalchemy import text

from app.audit import log_action
from app.db import SessionLocal, engine
from app.models import (
    Exam, ExamType,
    Building, Classroom, Course, CourseSection, Department, DepartmentMembership,
    DeliveryMode, EntryStatus, Lecturer, SemesterType, SessionType, Slot, User,
    UserRole, UserStatus, WeeklyScheduleEntry, Workgroup,
)
from app.normalize import normalize_lecturer_name
from app.security import hash_password

XLSX = r"C:\Users\ATLM\Desktop\staj_proje\CENG_2025_2026_Spring_Weekly_Schedule.xlsx"

SLOT_TIMES = [
    (1, "08:30", "09:15"), (2, "09:30", "10:15"), (3, "10:30", "11:15"),
    (4, "11:30", "12:15"), (5, "12:30", "13:15"), (6, "13:30", "14:15"),
    (7, "14:30", "15:15"), (8, "15:30", "16:15"), (9, "16:30", "17:15"),
]
SLOT_OF_START = {s: n for n, s, _ in SLOT_TIMES}
DAY_NO = {"MONDAY": 1, "TUESDAY": 2, "WEDNESDAY": 3, "THURSDAY": 4, "FRIDAY": 5}
ROMAN = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6}

# Hoca sütununda geçen ama KİŞİ OLMAYAN işaretler
NOT_A_LECTURER = {"synchronized - online", "görevlendirme", "online", ""}
DEFAULT_BUILDING = "Mühendislik Fakültesi"

TABLES = (
    "workgroups, users, invitation_tokens, departments, department_memberships, "
    "lecturers, buildings, classrooms, courses, course_sections, "
    "weekly_schedule_entries, exams, exam_classrooms, audit_logs, slots"
)


# ==================================================================
# Ayrıştırma
# ==================================================================

HEAD_RE = re.compile(
    r"^\s*([A-ZÇĞİÖŞÜ]{2,6})\s*(\d{3,4})"           # CENG 1004 / PHYS 1851
    r"(?:\s*[-(]\s*Section\s+([IVX]+|\d+)([^)\n]*)\)?)?",
    re.IGNORECASE,
)


def parse_head(text_):
    """Başlık satırından kod / şube / oturum türünü çıkarır."""
    m = HEAD_RE.match(text_.strip())
    if not m:
        return None
    alpha, num, sec, extra = m.groups()
    kind = SessionType.THEORY
    low = (extra or "").lower()
    if "lab" in low:
        kind = SessionType.LAB
    elif "uyg" in low or "practice" in low:
        kind = SessionType.PRACTICE
    if sec is None:
        section_no = 1
    elif sec.isdigit():
        section_no = int(sec)
    else:
        section_no = ROMAN.get(sec.upper(), 1)
    return {"code": f"{alpha.upper()}{num}", "section_no": section_no,
            "session_type": kind}


def room_of_head(text_):
    """Yıl sayfalarında derslik başlığın SONUNDAKİ parantezdedir: '... (FB07)'."""
    parens = re.findall(r"\(([^)]*)\)", text_)
    for p in reversed(parens):
        if "section" not in p.lower():
            return p.strip()
    return None


def day_columns(ws, header_row=2):
    """Gün → o güne ait sütunlar (birleşik başlık hücresi kadar geniş)."""
    spans = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row == header_row and rng.max_row == header_row:
            spans[rng.min_col] = list(range(rng.min_col, rng.max_col + 1))
    out = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip().upper() in DAY_NO:
            out[DAY_NO[v.strip().upper()]] = spans.get(c, [c])
    return out


def slot_rows(ws, time_col):
    """Satır → slot no (yalnız blok başlangıçları)."""
    out = {}
    for r in range(2, min(ws.max_row, 60) + 1):
        t = ws.cell(r, time_col).value
        if isinstance(t, str) and "-" in t:
            start = t.split("-")[0].strip()
            if start in SLOT_OF_START:
                out[r] = SLOT_OF_START[start]
    return out


def parse_year_sheet(ws, year):
    """Yıl sayfası: blok = 3 satır (kod+derslik / ad / hoca)."""
    rows, days = slot_rows(ws, 2), day_columns(ws)
    raw = []
    for day, cols in days.items():
        for col in cols:
            for r, slot in rows.items():
                head = ws.cell(r, col).value
                if not isinstance(head, str) or not head.strip():
                    continue
                info = parse_head(head)
                if not info:
                    continue
                name = ws.cell(r + 1, col).value
                lect = ws.cell(r + 2, col).value
                raw.append({**info, "year": year, "day": day, "slot": slot, "col": col,
                            "room": room_of_head(head),
                            "name": clean_name(name), "lecturer": clean(lect)})
    return raw


def parse_common_sheet(ws):
    """Ortak dersler: tek hücrede satırlar → kod / ad / hoca / derslik.
    Yıl, ders kodunun ilk rakamından türetilir (PHYS 1851 → 1. sınıf)."""
    rows, days = slot_rows(ws, 2), day_columns(ws)
    raw = []
    for day, cols in days.items():
        for col in cols:
            for r, slot in rows.items():
                cell = ws.cell(r, col).value
                if not isinstance(cell, str) or not cell.strip():
                    continue
                lines = [l.strip() for l in cell.split("\n") if l.strip()]
                if not lines:
                    continue
                info = parse_head(lines[0])
                if not info:
                    continue
                # 3. satırda hem hoca hem derslik olabilir (çok boşlukla ayrılmış)
                lect_line = lines[2] if len(lines) > 2 else ""
                room = lines[3] if len(lines) > 3 else None
                if room is None and re.search(r"\s{3,}", lect_line):
                    lect_line, room = re.split(r"\s{3,}", lect_line, 1)
                digit = re.search(r"\d", info["code"])
                year = int(info["code"][digit.start()]) if digit else 1
                raw.append({**info, "year": min(year, 4), "day": day, "slot": slot,
                            "col": col, "room": clean(room),
                            "name": clean_name(lines[1] if len(lines) > 1 else ""),
                            "lecturer": clean(lect_line)})
    return raw


def clean(v):
    return re.sub(r"\s+", " ", str(v)).strip() if isinstance(v, str) else ""


def clean_name(v):
    """Ders adından '(Section II)' gibi şube ekini atar — ad ders düzeyindedir."""
    return re.sub(r"\s*\(Section[^)]*\)", "", clean(v), flags=re.IGNORECASE)


def dedupe(raw):
    """Aynı yerleşim iki sayfada birden geçebiliyor.

    ENG3802 ve ATB3802 gibi ortak dersler hem "3rd Year" hem "Common Courses"
    sayfasında yazılı. Temizlenmezse aynı oturum iki kez yüklenir ve motor
    dersi KENDİSİYLE çakışıyor gösterir (sahte W2/W5).
    """
    seen, out = set(), []
    for e in raw:
        key = (e["code"], e["year"], e["section_no"], e["day"], e["slot"],
               e["session_type"])
        if key in seen:
            continue
        seen.add(key)
        out.append(e)
    return out


def merge_consecutive(raw):
    """Ardışık slotlardaki aynı yerleşimi tek girişe toplar (slot_count)."""
    raw.sort(key=lambda x: (x["year"], x["day"], x["col"], x["code"],
                            x["section_no"], x["slot"]))
    out = []
    for e in raw:
        p = out[-1] if out else None
        same = (p and all(p[k] == e[k] for k in
                          ("year", "day", "col", "code", "section_no",
                           "session_type", "room"))
                and p["slot"] + p["slot_count"] == e["slot"])
        if same:
            p["slot_count"] += 1
        else:
            out.append({**e, "slot_count": 1})
    return out


def is_elective(code):
    """ÇIKARIM (dosyada seçmelilik bilgisi YOK): CENG35xx = teknik seçmeli.

    Kanıt: 3. sınıfın CENG35xx dersleri aynı gün/saatte üst üste duruyor —
    zorunlu sayılsalardı motor haklı olarak W3 HARD üretirdi; oysa bunlar
    seçmeli havuzu, öğrenci birini seçer (K-05 → W4 uyarısı doğru davranış).
    Gerçek müfredat bilgisi geldiğinde bu satır kalkar.
    """
    return bool(re.match(r"^[A-ZÇĞİÖŞÜ]+\d5\d\d$", code))


def is_online(entry):
    """Online işareti hem hoca hem derslik hücresinde geçebiliyor (kaynak tutarsız)."""
    blob = f"{entry.get('lecturer') or ''} {entry.get('room') or ''}".lower()
    return "online" in blob or "synchronized" in blob


# Kaynak dosyada aynı bina farklı yazılmış: "Turizm Fak" / "Turizm Fakültesi".
# K-18'in yönetilen bina tablosu tam da bunun için var — içeri alırken tekilleştiriyoruz.
# NOT: "İİF" ile "İİBF" BİRLEŞTİRİLMEZ. Aynı bina olabilirler ama emin değiliz;
# birleştirmek iki dersi aynı odaya sokup GERÇEKTE OLMAYAN bir çakışma uydurdu.
# Şüpheli eşleme, kaçırılan eşlemeden daha zararlı — ayrı bırakıyoruz.
BUILDING_ALIASES: dict[str, str] = {}


def canon_building(name):
    name = re.sub(r"\bFak\.?$", "Fakültesi", clean(name))
    return BUILDING_ALIASES.get(name, name)


def split_room(s):
    """'Turizm Fakültesi - 206' → (bina, oda). Bina yoksa varsayılana düşer."""
    if not s:
        return None
    s = clean(s)
    m = re.split(r"\s*-\s+|\s+-\s*", s, 1)
    if len(m) == 2 and len(m[0]) > 3:
        return m[0].rstrip("-. "), m[1].strip()
    m2 = re.match(r"^(.*?(?:Fak|Fakültesi)\.?)\s+(.+)$", s, re.IGNORECASE)
    if m2:
        return m2.group(1).strip(), m2.group(2).strip()
    if " " in s and s.split(" ", 1)[0].isupper() and len(s.split(" ", 1)[0]) >= 3:
        a, b = s.split(" ", 1)
        return a, b
    return DEFAULT_BUILDING, s


# ==================================================================
# Yükleme
# ==================================================================

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    raw = []
    for sheet, yr in [("1st Year", 1), ("2nd Year", 2), ("3rd Year", 3), ("4th Year", 4)]:
        raw += parse_year_sheet(wb[sheet], yr)
    raw += parse_common_sheet(wb["Common Courses"])
    entries = merge_consecutive(dedupe(raw))
    print(f"• Dosyadan {len(entries)} yerleşim ayrıştırıldı.")

    with engine.begin() as conn:
        conn.execute(text(f"TRUNCATE {TABLES} RESTART IDENTITY CASCADE"))
    print("• Veritabanı sıfırlandı.")

    db = SessionLocal()
    for no, s, e in SLOT_TIMES:
        from datetime import time as _t
        hh, mm = map(int, s.split(":")); eh, em = map(int, e.split(":"))
        db.add(Slot(slot_no=no, start_time=_t(hh, mm), end_time=_t(eh, em)))

    wg = Workgroup(name="Mühendislik Fakültesi",
                   allowed_email_domain="muh.example.edu.tr")
    db.add(wg); db.flush()
    admin = User(workgroup_id=wg.id, name="Fakülte Yöneticisi",
                 email="admin@muh.example.edu.tr",
                 password_hash=hash_password("admin1234"),
                 role=UserRole.ADMIN, status=UserStatus.ACTIVE)
    db.add(admin); db.flush()
    wg.created_by = admin.id

    dep = Department(workgroup_id=wg.id, name="Bilgisayar Mühendisliği", code="BM")
    db.add(dep); db.flush()
    log_action(db, admin, "CREATE", "department", dep.id, dep)

    # --- binalar + derslikler ---
    buildings, rooms = {}, {}

    def get_room(raw_room):
        if not raw_room:
            return None
        b_name, r_code = split_room(raw_room)
        b_name = canon_building(b_name)
        if b_name not in buildings:
            b = Building(workgroup_id=wg.id, name=b_name)
            db.add(b); db.flush()
            log_action(db, admin, "CREATE", "building", b.id, b)
            buildings[b_name] = b
        key = (b_name, r_code)
        if key not in rooms:
            # SENTETİK: kapasiteler dosyada yok — sınıf 60 / amfi 150 varsayıldı
            cap = 150 if re.search(r"amfi", r_code, re.IGNORECASE) else 60
            c = Classroom(workgroup_id=wg.id, building_id=buildings[b_name].id,
                          room_code=r_code, capacity=cap, exam_capacity=cap // 2)
            db.add(c); db.flush()
            log_action(db, admin, "CREATE", "classroom", c.id, c)
            rooms[key] = c
        return rooms[key]

    # --- hocalar (K-08: normalized_name ile tekilleşir) ---
    lecturers = {}

    def get_lecturer(full_name, code):
        # Hocası yazılmayan dersler ("Görevlendirme", online) DERS BAŞINA ayrı
        # yer tutucu alır. Hepsi tek "Belirtilmemiş"te toplansaydı motor onları
        # aynı kişi sanıp sahte W2 (hoca çakışması) üretirdi — nitekim üretti.
        name = (full_name if full_name.lower() not in NOT_A_LECTURER
                else f"Belirtilmemiş ({code})")
        norm = normalize_lecturer_name(name)
        if norm not in lecturers:
            l = Lecturer(workgroup_id=wg.id, full_name=name, normalized_name=norm,
                         source="IMPORT")
            db.add(l); db.flush()
            log_action(db, admin, "CREATE", "lecturer", l.id, l)
            lecturers[norm] = l
        return lecturers[norm]

    # --- T+U+L: dosyada yok, YERLEŞEN slotlardan türetilir (W8 gürültüsü olmasın) ---
    hours = defaultdict(lambda: defaultdict(int))     # (code,year) -> {tür: slot}
    per_section = defaultdict(lambda: defaultdict(int))
    for e in entries:
        per_section[(e["code"], e["year"], e["section_no"])][e["session_type"]] += e["slot_count"]
    for (code, year, _sec), by_type in per_section.items():
        for st, n in by_type.items():
            hours[(code, year)][st] = max(hours[(code, year)][st], n)

    courses, sections = {}, {}
    for e in entries:
        ckey = (e["code"], e["year"])
        if ckey not in courses:
            h = hours[ckey]
            c = Course(department_id=dep.id, year=e["year"],
                       semester=SemesterType.SPRING, code=e["code"],
                       name=e["name"] or e["code"], is_elective=is_elective(e["code"]),
                       hours_theory=h.get(SessionType.THEORY, 0),
                       hours_practice=h.get(SessionType.PRACTICE, 0),
                       hours_lab=h.get(SessionType.LAB, 0))
            db.add(c); db.flush()
            log_action(db, admin, "CREATE", "course", c.id, c)
            courses[ckey] = c
        skey = (*ckey, e["section_no"])
        if skey not in sections:
            s = CourseSection(course_id=courses[ckey].id, section_no=e["section_no"],
                              lecturer_id=get_lecturer(e["lecturer"], e["code"]).id,
                              expected_students=45)   # SENTETİK: dosyada yok
            db.add(s); db.flush()
            log_action(db, admin, "CREATE", "course_section", s.id, s)
            sections[skey] = s

    # --- haftalık yerleşimler ---
    # Şubenin VARSAYILAN dersliği: dosyada ayrı bir alan yok, yerleşimlerinde
    # en çok kullandığı derslikten türetiyoruz. Boş bırakılsaydı Dersler ekranı
    # her şube için "derslik yok" derdi — oysa bilgi elimizde.
    section_rooms = defaultdict(list)

    for e in entries:
        online = is_online(e)
        room = None if online else get_room(e["room"])
        if room is not None:
            section_rooms[(e["code"], e["year"], e["section_no"])].append(room.id)
        w = WeeklyScheduleEntry(
            section_id=sections[(e["code"], e["year"], e["section_no"])].id,
            classroom_id=room.id if room else None,
            day_of_week=e["day"], start_slot=e["slot"],
            slot_count=min(e["slot_count"], 9 - e["slot"] + 1),
            session_type=e["session_type"],
            delivery_mode=DeliveryMode.ONLINE_SYNC if online else DeliveryMode.FACE_TO_FACE, created_by=admin.id)
        db.add(w); db.flush()
        log_action(db, admin, "CREATE", "weekly_entry", w.id, w)

    # Şubenin en sık kullandığı derslik = varsayılan derslik
    for skey, room_ids in section_rooms.items():
        sections[skey].default_classroom_id = Counter(room_ids).most_common(1)[0][0]

    # ==============================================================
    # TEST KURGUSU — dosyadan DEĞİL, elle eklenen sentetik veri.
    #
    # Amaç: alt hesabın çakışma görünürlüğünü sınayabilmek. İçe aktarılan CENG
    # programı tek bölümlük olduğu için "başka bölümle çakışma" ve "beni
    # ilgilendirmeyen çakışma" durumları o veriyle üretilemiyor.
    #
    # Kurulan üç durum:
    #   1) MM101 ↔ bir BM dersi: AYNI derslik/saat  → bölümler ARASI çakışma
    #      (BM alt hesabı GÖRMELİ — çözebilmesi için karşı tarafı bilmeli)
    #   2) MM201 ↔ MM202: kendi odalarında çakışma  → yalnız MM'i ilgilendirir
    #      (BM alt hesabı GÖRMEMELİ)
    #   3) BM'ye atanmış bir alt hesap (yazma yetkileriyle)
    # ==============================================================
    mm = Department(workgroup_id=wg.id, name="Makine Mühendisliği", code="MM")
    db.add(mm); db.flush()
    log_action(db, admin, "CREATE", "department", mm.id, mm)

    mm_bina = Building(workgroup_id=wg.id, name="Makine Binası")
    db.add(mm_bina); db.flush()
    mm_oda = Classroom(workgroup_id=wg.id, building_id=mm_bina.id, room_code="M-101",
                       capacity=60, exam_capacity=30)
    db.add(mm_oda); db.flush()

    def mm_ders(kod, ad, hoca_ad, gun, slot, oda_id, saat=2):
        c = Course(department_id=mm.id, year=1, semester=SemesterType.SPRING,
                   code=kod, name=ad, is_elective=False,
                   hours_theory=saat, hours_practice=0, hours_lab=0)
        db.add(c); db.flush()
        s = CourseSection(course_id=c.id, section_no=1,
                          lecturer_id=get_lecturer(hoca_ad, kod).id,
                          expected_students=40, default_classroom_id=oda_id)
        db.add(s); db.flush()
        w = WeeklyScheduleEntry(section_id=s.id, classroom_id=oda_id, day_of_week=gun,
                                start_slot=slot, slot_count=saat,
                                session_type=SessionType.THEORY,
                                delivery_mode=DeliveryMode.FACE_TO_FACE, created_by=admin.id)
        db.add(w); db.flush()
        log_action(db, admin, "CREATE", "weekly_entry", w.id, w)

    # (1) Bölümler arası: BM'nin dersliği olan ilk girişinin yerine MM dersi koy
    hedef = next((e for e in db.query(WeeklyScheduleEntry)
                  .filter(WeeklyScheduleEntry.classroom_id.isnot(None))
                  .order_by(WeeklyScheduleEntry.id).all()), None)
    if hedef is not None:
        mm_ders("MM101", "Statik", "Dr. Kerem Aydın",
                hedef.day_of_week, hedef.start_slot, hedef.classroom_id)

    # (2) MM'in kendi içinde: iki ders aynı odada aynı saatte
    mm_ders("MM201", "Termodinamik", "Dr. Sinan Yılmaz", 3, 2, mm_oda.id)
    mm_ders("MM202", "Akışkanlar Mekaniği", "Dr. Pelin Kara", 3, 2, mm_oda.id)

    # ---- SINAVLAR (dosyada yok; sınav takvimi ekranını beslemek için) ----
    # Gerçek final haftası deseni: 15-19 Haziran 2026 (Pzt-Cum), günde birkaç
    # sınav, farklı saatler. K-06: saat penceresi yok, 18:00 sınavı geçerli.
    sinav_plani = [
        # (ders kodu, yıl, tip, tarih, saat, süre, derslik sayısı)
        ("CENG1004", 1, ExamType.FINAL, date(2026, 6, 15), time(9, 0), 90, 2),
        ("CENG1008", 1, ExamType.FINAL, date(2026, 6, 15), time(13, 0), 90, 1),
        ("MATH1852", 1, ExamType.FINAL, date(2026, 6, 16), time(9, 0), 120, 2),
        ("PHYS1851", 1, ExamType.FINAL, date(2026, 6, 16), time(14, 0), 90, 1),
        ("CENG2010", 2, ExamType.FINAL, date(2026, 6, 17), time(9, 0), 90, 1),
        ("CENG2014", 2, ExamType.FINAL, date(2026, 6, 17), time(13, 0), 90, 1),
        ("CENG2032", 2, ExamType.FINAL, date(2026, 6, 18), time(9, 0), 90, 1),
        # Bilerek ÇAKIŞTIRILAN ikili: aynı gün/saat, aynı derslik → E1 HARD
        ("CENG3004", 3, ExamType.FINAL, date(2026, 6, 18), time(14, 0), 90, 1),
        ("CENG3512", 3, ExamType.FINAL, date(2026, 6, 18), time(14, 0), 90, 1),
        # Akşam sınavı: saat kısıtı olmadığını gösterir (K-06)
        ("CENG4012", 4, ExamType.FINAL, date(2026, 6, 19), time(18, 0), 120, 1),
    ]
    oda_havuzu = [r for (_, _), r in sorted(rooms.items(), key=lambda kv: kv[1].id)]
    for kod, yil, tip, tarih, saat, sure, oda_adet in sinav_plani:
        c = courses.get((kod, yil))
        if c is None:
            continue
        sec = next((s for (ck, cy, _sn), s in sections.items()
                    if ck == kod and cy == yil), None)
        x = Exam(course_id=c.id, exam_type=tip, exam_date=tarih, start_time=saat,
                 duration_minutes=sure,
                 lecturer_id=sec.lecturer_id if sec else admin.id, created_by=admin.id)
        # Çakışan ikili AYNI dersliği paylaşsın diye havuzun başından veriyoruz
        x.classrooms = oda_havuzu[:oda_adet]
        db.add(x); db.flush()
        log_action(db, admin, "CREATE", "exam", x.id, x)

    # (3) BM'ye atanmış alt hesap
    alt = User(workgroup_id=wg.id, name="Zeynep Nur",
               email="zeynep@muh.example.edu.tr",
               password_hash=hash_password("althesap123"),
               role=UserRole.SUB_ACCOUNT, status=UserStatus.ACTIVE,
               can_manage_courses=True, can_manage_weekly=True, can_manage_exams=True)
    db.add(alt); db.flush()
    db.add(DepartmentMembership(user_id=alt.id, department_id=dep.id))
    log_action(db, admin, "INVITE", "user", alt.id, alt)
    log_action(db, alt, "ACTIVATE", "user", alt.id, alt)

    db.commit()
    print(f"• Yüklendi: {len(courses)} ders · {len(sections)} şube · "
          f"{len(lecturers)} hoca · {len(rooms)} derslik · {len(entries)} yerleşim")
    db.close()
    print("""
┌──────────────────────────────────────────────────────────────┐
│  Giriş: admin@muh.example.edu.tr / admin1234                 │
│  Haftalık Program'da DÖNEM = "Bahar" seç (dosya bahar dönemi)│
└──────────────────────────────────────────────────────────────┘""")


if __name__ == "__main__":
    main()
